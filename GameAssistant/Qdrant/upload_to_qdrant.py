# ============================================================
# 杀戮尖塔2 游戏数据向量化 → Qdrant
# 读取 neo4j/ 下的 Excel 表格，向量化后批量入库
# ============================================================
# 安装依赖：
#   pip install qdrant-client openai openpyxl pandas
# ============================================================

import os, uuid, time
import openpyxl
from dotenv import load_dotenv
from openai import OpenAI
from qdrant_client import QdrantClient, models

# ── 配置（从 .env 文件读取）────────────────────────────────
load_dotenv(os.path.join(os.path.dirname(__file__), "..", ".env"))
QDRANT_URL     = os.getenv("QDRANT_URL")
QDRANT_API_KEY = os.getenv("QDRANT_API_KEY")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")

COLLECTION_NAME = "STS2_GameData"
EMBED_MODEL     = "text-embedding-3-small"   # 1536 维
EMBED_DIM       = 1536
BATCH_SIZE      = 5    # 每批上传数量（小批次避免超时）

XLSX_DIR = os.path.join(os.path.dirname(__file__), "..", "neo4j")
# ──────────────────────────────────────────────────────────

openai_client = OpenAI(api_key=OPENAI_API_KEY)
qdrant_client = QdrantClient(url=QDRANT_URL, api_key=QDRANT_API_KEY, timeout=60)


# ── 文本构建规则（每张表 → 可搜索的句子）─────────────────
def build_text(table: str, row: dict) -> str:
    if table == "角色":
        return (f"角色: {row.get('name','')}（{row.get('name_en','')}）"
                f" | 牌色: {row.get('card_color','')} | 核心机制: {row.get('special_mechanic','')}"
                f" | 初始遗物: {row.get('starting_relic','')} 效果: {row.get('relic_effect','')}"
                f" | 简介: {row.get('description','')}")

    elif table == "遗物":
        return (f"遗物: {row.get('name','')} | 归属角色: {row.get('character_id','')}"
                f" | 等级: {row.get('tier','')} | 效果: {row.get('description','')}"
                f" | 风味: {row.get('flavor','')}")

    elif table == "卡牌":
        return (f"卡牌: {row.get('name','')} | 角色: {row.get('character_id','')}"
                f" | 稀有度: {row.get('rarity','')} | 类型: {row.get('type','')}"
                f" | 费用: {row.get('cost','')} | 效果: {row.get('description','')}")

    elif table == "药水":
        return (f"药水: {row.get('name','')} | 颜色: {row.get('color','')}"
                f" | 等级: {row.get('tier','')} | 效果: {row.get('description','')}")

    elif table == "关系_角色_遗物":
        return (f"关系: 角色 {row.get('start_id(Character)','')} 的初始遗物是"
                f" {row.get('end_id(Relic)','')} | 备注: {row.get('notes','')}")

    elif table == "关系_怪物_区域":
        return (f"关系: 怪物 {row.get('monster_name','')} 出现在区域"
                f" {row.get('zone_name','')} | 类型: {row.get('monster_type','')}")

    elif table == "关系_卡牌_升级":
        return (f"关系: 卡牌 {row.get('base_name','')} 升级后变为"
                f" {row.get('upgrade_name','')}")

    elif table == "关系_角色_卡牌":
        return (f"关系: 角色 {row.get('character_name','')} 拥有卡牌"
                f" {row.get('card_name','')} | 稀有度: {row.get('card_rarity','')}")

    return str(row)


# ── 图片列（不向量化，但保留在 payload）─────────────────
IMAGE_COLS = {"角色立绘", "图标", "image_file", "image"}


# ── 读取 Excel → list[dict] ────────────────────────────
def load_xlsx(path: str) -> tuple[str, list[dict]]:
    """返回 (sheet_name, rows)，自动跳过图片占位列"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {}
        for h, v in zip(headers, row):
            if h in IMAGE_COLS:
                continue   # 跳过图片列（无文本值）
            d[h] = str(v) if v is not None else ""
        rows.append(d)
    wb.close()
    return ws.title, rows


# ── OpenAI 批量 Embedding ──────────────────────────────
def get_embeddings(texts: list[str]) -> list[list[float]]:
    resp = openai_client.embeddings.create(input=texts, model=EMBED_MODEL)
    return [item.embedding for item in resp.data]


# ── 主流程 ─────────────────────────────────────────────
def main():
    # 1. 建或重建 Collection
    existing = [c.name for c in qdrant_client.get_collections().collections]
    if COLLECTION_NAME in existing:
        print(f"Collection '{COLLECTION_NAME}' 已存在，跳过创建")
    else:
        qdrant_client.create_collection(
            collection_name=COLLECTION_NAME,
            vectors_config=models.VectorParams(
                size=EMBED_DIM,
                distance=models.Distance.COSINE,
            ),
        )
        print(f"✓ 创建 Collection: {COLLECTION_NAME}")

    # 2. 文件 → 表名 映射
    file_table_map = {
        "01_节点_角色.xlsx":          "角色",
        "02_节点_遗物.xlsx":          "遗物",
        "03_节点_卡牌.xlsx":          "卡牌",
        "04_节点_药水.xlsx":          "药水",
        "05_关系_角色_初始遗物.xlsx": "关系_角色_遗物",
        "06_关系_怪物_区域.xlsx":     "关系_怪物_区域",
        "09_关系_卡牌_升级.xlsx":     "关系_卡牌_升级",
        "10_关系_角色_卡牌.xlsx":     "关系_角色_卡牌",
    }

    total_uploaded = 0

    for filename, table_name in file_table_map.items():
        path = os.path.join(XLSX_DIR, filename)
        if not os.path.exists(path):
            print(f"  ⚠ 文件不存在，跳过: {filename}")
            continue

        _, rows = load_xlsx(path)
        print(f"\n[{table_name}] {len(rows)} 行，开始向量化...")

        # 分批处理
        for i in range(0, len(rows), BATCH_SIZE):
            batch_rows = rows[i:i + BATCH_SIZE]
            texts = [build_text(table_name, r) for r in batch_rows]

            embeddings = get_embeddings(texts)

            points = []
            for row, emb, text in zip(batch_rows, embeddings, texts):
                points.append(models.PointStruct(
                    id=str(uuid.uuid4()),
                    vector=emb,
                    payload={
                        "table":   table_name,
                        "text":    text,      # 存入原始文本，方便 RAG 召回
                        **row               # 所有结构化字段
                    }
                ))

            for attempt in range(3):
                try:
                    qdrant_client.upsert(collection_name=COLLECTION_NAME, points=points)
                    break
                except Exception as e:
                    if attempt == 2:
                        raise
                    print(f"  重试 {attempt+1}/3 ({e.__class__.__name__})", flush=True)
                    time.sleep(2)
            total_uploaded += len(points)
            print(f"  上传 {i+len(batch_rows)}/{len(rows)}", flush=True)
            time.sleep(0.3)

    print(f"\n{'='*50}")
    print(f"✓ 全部完成，共上传 {total_uploaded} 条向量")
    print(f"  Collection: {COLLECTION_NAME}")
    info = qdrant_client.get_collection(COLLECTION_NAME)
    print(f"  向量总数: {info.points_count}")


if __name__ == "__main__":
    main()
