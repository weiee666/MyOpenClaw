#!/usr/bin/env python3
# ============================================================
# Fix ASR (Whisper) entity name errors in transcripts
# Uses sliding-window edit distance to find misrecognized
# card/relic/character/potion names and apply corrections.
# ============================================================

import json
import os
import re

import openpyxl

BASE_DIR       = os.path.dirname(__file__)
TRANSCRIPT_DIR = os.path.join(BASE_DIR, "transcripts")
DOWNLOADS_DIR  = os.path.join(BASE_DIR, "downloads")
NEO4J_DIR      = os.path.join(BASE_DIR, "..", "neo4j")

EXCEL_MAP = {
    "characters": "01_节点_角色.xlsx",
    "relics":     "02_节点_遗物.xlsx",
    "cards":      "03_节点_卡牌.xlsx",
    "potions":    "04_节点_药水.xlsx",
}

# Edit distance thresholds by name length
# (shorter names need stricter matching to avoid false positives)
def max_dist(name: str) -> int:
    n = len(name)
    if n <= 2:
        return 0   # exact only for 1-2 char names
    if n <= 4:
        return 1
    if n <= 7:
        return 2
    return 3


def edit_distance(a: str, b: str) -> int:
    """Standard Levenshtein edit distance."""
    la, lb = len(a), len(b)
    dp = list(range(lb + 1))
    for i in range(1, la + 1):
        prev = dp[0]
        dp[0] = i
        for j in range(1, lb + 1):
            temp = dp[j]
            if a[i - 1] == b[j - 1]:
                dp[j] = prev
            else:
                dp[j] = 1 + min(prev, dp[j], dp[j - 1])
            prev = temp
    return dp[lb]


def load_names(xlsx_path: str) -> list[str]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(max_row=1))]
    ni = headers.index("name")
    names = [str(r[ni]).strip() for r in ws.iter_rows(min_row=2, values_only=True) if r[ni]]
    wb.close()
    return names


def load_all_entities() -> dict[str, list[str]]:
    result = {}
    for key, fname in EXCEL_MAP.items():
        result[key] = load_names(os.path.join(NEO4J_DIR, fname))
    return result


def find_corrections(text: str, entities: dict[str, list[str]]) -> list[dict]:
    """
    Slide a window of each entity name's length over text.
    Return list of {wrong, correct, category, positions} for near-matches.
    """
    corrections = []
    seen_wrong = set()

    all_names = []
    for cat, names in entities.items():
        for name in names:
            all_names.append((cat, name))

    # Sort by length desc so longer names are matched first
    all_names.sort(key=lambda x: -len(x[1]))

    for cat, name in all_names:
        n = len(name)
        threshold = max_dist(name)
        if threshold == 0:
            # Exact only — no fuzzy needed, skip
            continue

        # Find all windows of length n in text
        for i in range(len(text) - n + 1):
            window = text[i:i + n]
            if window == name:
                continue  # already correct
            if window in seen_wrong:
                continue  # already flagged

            dist = edit_distance(window, name)
            if 0 < dist <= threshold:
                # Make sure this window doesn't partially overlap an already-found correction
                # Gather context around position
                ctx_start = max(0, i - 8)
                ctx_end   = min(len(text), i + n + 8)
                context   = text[ctx_start:ctx_end]
                context   = context.replace(window, f"【{window}】")

                corrections.append({
                    "wrong":    window,
                    "correct":  name,
                    "category": cat,
                    "dist":     dist,
                    "context":  context,
                })
                seen_wrong.add(window)
                break  # one entry per wrong string

    return corrections


def apply_corrections(text: str, corrections: list[dict]) -> tuple[str, int]:
    """Replace wrong strings with correct ones. Returns (new_text, count)."""
    count = 0
    for c in corrections:
        occurrences = text.count(c["wrong"])
        if occurrences > 0:
            text = text.replace(c["wrong"], c["correct"])
            count += occurrences
    return text, count


def fix_video(video_id: str, entities: dict, confirmed: dict[str, str]) -> dict:
    """
    Apply confirmed corrections to _chunks.json and regenerate _enriched.json.
    confirmed: {wrong -> correct}
    """
    chunks_path   = os.path.join(TRANSCRIPT_DIR, f"{video_id}_chunks.json")
    enriched_path = os.path.join(TRANSCRIPT_DIR, f"{video_id}_enriched.json")

    with open(chunks_path, encoding="utf-8") as f:
        chunks = json.load(f)

    total_fixed = 0
    for chunk in chunks:
        original = chunk["text"]
        for wrong, correct in confirmed.items():
            if wrong in original:
                chunk["text"] = chunk["text"].replace(wrong, correct)
                total_fixed += original.count(wrong)

    with open(chunks_path, "w", encoding="utf-8") as f:
        json.dump(chunks, f, ensure_ascii=False, indent=2)

    # Regenerate enriched JSON
    with open(os.path.join(DOWNLOADS_DIR, f"{video_id}_meta.json"), encoding="utf-8") as f:
        meta = json.load(f)

    full_text = "".join(c["text"] for c in chunks)
    matched = {k: [n for n in names if n in full_text] for k, names in entities.items()}

    enriched = {
        "video_id":    video_id,
        "video_title": meta.get("title", ""),
        "video_url":   meta.get("url", ""),
        "full_text":   full_text,
        "characters":  matched["characters"],
        "cards":       matched["cards"],
        "relics":      matched["relics"],
        "potions":     matched["potions"],
    }
    with open(enriched_path, "w", encoding="utf-8") as f:
        json.dump(enriched, f, ensure_ascii=False, indent=2)

    print(f"[fix] {video_id}: {total_fixed} replacements applied")
    return enriched


def scan_all():
    """Scan all transcripts and print potential corrections for review."""
    entities = load_all_entities()

    video_ids = sorted(
        f.replace("_chunks.json", "")
        for f in os.listdir(TRANSCRIPT_DIR)
        if f.endswith("_chunks.json")
    )

    all_corrections = {}

    for vid in video_ids:
        with open(os.path.join(TRANSCRIPT_DIR, f"{vid}_chunks.json"), encoding="utf-8") as f:
            chunks = json.load(f)
        full_text = "".join(c["text"] for c in chunks)

        corrections = find_corrections(full_text, entities)
        if corrections:
            all_corrections[vid] = corrections
            meta_path = os.path.join(DOWNLOADS_DIR, f"{vid}_meta.json")
            with open(meta_path, encoding="utf-8") as f:
                title = json.load(f).get("title", vid)
            print(f"\n{'='*60}")
            print(f"视频: {title}")
            print(f"{'='*60}")
            for c in corrections:
                print(f"  [{c['category']}] 「{c['wrong']}」→「{c['correct']}」  编辑距离:{c['dist']}")
                print(f"    上下文: ...{c['context']}...")

    return all_corrections


# ── Manual correction table (verified by scan + Excel cross-check) ─────────────
# Format: {video_id: {wrong_asr_text: correct_entity_name}}
MANUAL_CORRECTIONS: dict[str, dict[str, str]] = {
    # 《战士新手攻略》
    "BV12JX5BnE2r": {
        "无情冷攻": "无情猛攻",   # 猛→冷 音近误
        "溃然不动": "岿然不动",   # 溃→岿 形近误
        "剑鸣打击": "剑柄打击",   # 鸣→柄 ASR误
    },
    # 《战士4种无限构筑》
    "BV15QwJzTEj4": {
        "飞红披风": "绯红披风",   # 飞→绯 同音误
    },
    # 《鸡煲新手攻略》
    "BV1BU9LBVEox": {
        "娇嫩绝草": "娇嫩蕨草",   # 绝→蕨 同音误
        "佩尔之兆": "佩尔之爪",   # 兆→爪 ASR误
        "变废为宝": "化废为宝",   # 变→化 近义误
        "螺旋钻戒": "螺旋钻击",   # 戒→击 ASR误
        "玻璃工业": "玻璃工艺",   # 业→艺 同音误
        "L速脱离": "高速脱离",    # L→高 ASR误（念字母而非汉字）
        "在利用":  "再利用",      # 在→再 同音误
        "双重释放": "四重释放",   # 双→四 同音误（鸡煲语境）
    },
    # 《猎手新手攻略》
    "BV1JhAgzXEKG": {
        "刀刀陷阱": "刀刃陷阱",   # 刀→刃 重复误
        "灵动步伐": "灵动步法",   # 伐→法 同音误
        "反能反应": "本能反应",   # 反能→本能 ASR误
        "触摸可及": "触不可及",   # 摸→不 ASR误
        "欧刃之舞": "刀刃之舞",   # 欧→刀 ASR误
        "翻越撑街": "翻越撑击",   # 街→击 同音误
    },
    # 《猎手+储君无限构筑》
    "BV1ZJwrzqEoe": {
        "战术大师": "战略大师",   # 术→略 音近误
        "星微序列": "星位序列",   # 微→位 同音误
    },
    # 《鸡煲3种无限构筑》
    "BV1u7wQzSE4v": {
        "双重释放": "四重释放",   # 双→四 同音误（鸡煲语境）
        "火箭飞船": "火箭飞拳",   # 船→拳 ASR误
        "在利用":  "再利用",      # 在→再 同音误
    },
    # 《亡灵3种无限构筑》
    "BV1vxwTziE3G": {
        "削肉戏法": "血肉戏法",   # 削→血 ASR误
        "预戒时间": "预借时间",   # 戒→借 同音误
        "死亡行径": "死亡行军",   # 径→军 ASR误
    },
}


def apply_manual_corrections(dry_run: bool = False):
    """Apply MANUAL_CORRECTIONS to chunks and regenerate enriched files."""
    entities = load_all_entities()

    for vid, corrections_map in MANUAL_CORRECTIONS.items():
        # Only keep corrections where wrong != correct
        confirmed = {k: v for k, v in corrections_map.items() if k != v}
        if not confirmed:
            print(f"[fix] {vid}: no corrections to apply")
            continue

        chunks_path = os.path.join(TRANSCRIPT_DIR, f"{vid}_chunks.json")
        if not os.path.exists(chunks_path):
            print(f"[fix] {vid}: chunks file not found, skipping")
            continue

        if dry_run:
            with open(chunks_path, encoding="utf-8") as f:
                chunks = json.load(f)
            full_text = "".join(c["text"] for c in chunks)
            for wrong, correct in confirmed.items():
                count = full_text.count(wrong)
                if count:
                    print(f"  [DRY] {vid}: 「{wrong}」→「{correct}」 ({count}处)")
        else:
            fix_video(vid, entities, confirmed)


if __name__ == "__main__":
    import sys

    if "--scan" in sys.argv:
        print("扫描模式：查找所有视频中的潜在实体名称错误...\n")
        scan_all()
    elif "--dry" in sys.argv:
        print("预览模式：以下更改将被应用：\n")
        apply_manual_corrections(dry_run=True)
    else:
        print("应用人工确认的修正...\n")
        apply_manual_corrections(dry_run=False)
        print("\n完成。请重新运行 inject_qdrant.py 更新 Qdrant。")
