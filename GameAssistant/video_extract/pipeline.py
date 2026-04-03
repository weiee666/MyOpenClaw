#!/usr/bin/env python3
# ============================================================
# Bilibili Video → Whisper → Qdrant pipeline
# 单条: python pipeline.py --url <bilibili_url>
# 批量: python pipeline.py --excel videos.xlsx
# ============================================================

import argparse
import os
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(__file__))

from download_video     import download_audio
from transcribe_whisper import transcribe_and_chunk
from enrich_transcript  import enrich_video, load_all_entities
from inject_qdrant      import inject_chunks, inject_video, COLLECTION_NAME

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "videos.xlsx")


def run_pipeline(
    url:        str,
    collection: str = COLLECTION_NAME,
    entities:   dict = None,
):
    print("=" * 60)
    print(f"处理: {url[:60]}...")
    print("=" * 60)

    # ── Step 1: Download ────────────────────────────────────
    print("\n[Step 1/3] 下载音频...")
    metadata    = download_audio(url)
    video_id    = metadata["video_id"]
    video_title = metadata["title"]
    audio_path  = metadata["audio_path"]
    print(f"  video_id : {video_id}")
    print(f"  title    : {video_title}")

    # ── Step 2: Transcribe ──────────────────────────────────
    print(f"\n[Step 2/3] 语音转文字...")
    chunks = transcribe_and_chunk(
        audio_path=audio_path,
        video_id=video_id,
    )
    print(f"  生成 {len(chunks)} 个片段")

    # ── Step 2.5: Enrich ────────────────────────────────────
    print(f"\n[Step 2.5/3] 富化转录（实体匹配）...")
    if entities is None:
        entities = load_all_entities()
    enriched = enrich_video(video_id, entities)
    print(f"  角色={enriched['characters']}, 卡牌数={len(enriched['cards'])}, "
          f"遗物={enriched['relics']}, 药水={enriched['potions']}")

    # ── Step 3: Inject ──────────────────────────────────────
    print(f"\n[Step 3/3] 写入 Qdrant ({collection})...")
    inject_video(enriched, collection=collection)

    print(f"\n✓ 完成: {video_title} | 1 条向量")
    return True


def run_from_excel(
    excel_path: str = EXCEL_PATH,
    collection: str = COLLECTION_NAME,
):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # 找列索引（从1开始）
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    try:
        url_col  = headers.index("url") + 1
        done_col = headers.index("is_done") + 1
    except ValueError as e:
        print(f"Excel 列名错误: {e}")
        print(f"当前列名: {headers}")
        sys.exit(1)

    # 预加载实体词表，所有视频共用
    entities = load_all_entities()

    total_rows = ws.max_row - 1
    skipped = 0
    success = 0
    failed  = 0

    print(f"Excel: {excel_path}，共 {total_rows} 条")
    print("=" * 60)

    for row_idx in range(2, ws.max_row + 1):
        url     = ws.cell(row=row_idx, column=url_col).value
        is_done = ws.cell(row=row_idx, column=done_col).value

        if not url:
            continue

        if is_done == 1:
            print(f"[{row_idx-1}/{total_rows}] 跳过（已完成）: {url[:60]}...")
            skipped += 1
            continue

        print(f"\n[{row_idx-1}/{total_rows}] 开始处理...")
        try:
            run_pipeline(url=url, collection=collection, entities=entities)
            # 标记完成
            ws.cell(row=row_idx, column=done_col).value = 1
            wb.save(excel_path)
            success += 1
        except Exception as e:
            print(f"  ✗ 失败: {e}")
            failed += 1

    print("\n" + "=" * 60)
    print(f"批量处理完成: 成功 {success} | 跳过 {skipped} | 失败 {failed}")


def main():
    parser = argparse.ArgumentParser(
        description="Bilibili → Whisper → Qdrant"
    )
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--url",   help="单条 Bilibili 视频 URL")
    group.add_argument("--excel", nargs="?", const=EXCEL_PATH,
                       help=f"批量模式，读取 Excel 文件（默认: videos.xlsx）")

    parser.add_argument("--collection", default=COLLECTION_NAME,
                        help=f"Qdrant 集合名（默认: {COLLECTION_NAME}）")
    args = parser.parse_args()

    if args.url:
        run_pipeline(
            url=args.url,
            collection=args.collection,
        )
    else:
        run_from_excel(
            excel_path=args.excel,
            collection=args.collection,
        )


if __name__ == "__main__":
    main()
