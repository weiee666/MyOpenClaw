#!/usr/bin/env python3
# ============================================================
# Step 2.5: Enrich transcript with entity matching from Excel
# Input:  transcripts/{video_id}_chunks.json
#         downloads/{video_id}_meta.json
# Output: transcripts/{video_id}_enriched.json
# ============================================================

import json
import os

import openpyxl

BASE_DIR      = os.path.dirname(__file__)
TRANSCRIPT_DIR = os.path.join(BASE_DIR, "transcripts")
DOWNLOADS_DIR  = os.path.join(BASE_DIR, "downloads")
NEO4J_DIR      = os.path.join(BASE_DIR, "..", "neo4j")

EXCEL_MAP = {
    "characters": "01_节点_角色.xlsx",
    "relics":     "02_节点_遗物.xlsx",
    "cards":      "03_节点_卡牌.xlsx",
    "potions":    "04_节点_药水.xlsx",
}


def _load_names(xlsx_path: str) -> list[str]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(cell.value).strip() if cell.value else "" for cell in next(ws.iter_rows(max_row=1))]
    try:
        name_col = headers.index("name")
    except ValueError:
        raise ValueError(f"'name' column not found in {xlsx_path}. Headers: {headers}")

    names = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[name_col]
        if val and str(val).strip():
            names.append(str(val).strip())
    wb.close()
    return names


def load_all_entities() -> dict[str, list[str]]:
    result = {}
    for key, filename in EXCEL_MAP.items():
        path = os.path.join(NEO4J_DIR, filename)
        result[key] = _load_names(path)
    return result


def enrich_video(video_id: str, entities: dict[str, list[str]]) -> dict:
    enriched_path = os.path.join(TRANSCRIPT_DIR, f"{video_id}_enriched.json")
    if os.path.exists(enriched_path):
        print(f"[enrich] Skipping {video_id} (already exists)")
        with open(enriched_path, encoding="utf-8") as f:
            return json.load(f)

    # Load chunks
    chunks_path = os.path.join(TRANSCRIPT_DIR, f"{video_id}_chunks.json")
    with open(chunks_path, encoding="utf-8") as f:
        chunks = json.load(f)
    full_text = "".join(c["text"] for c in chunks)

    # Load meta
    meta_path = os.path.join(DOWNLOADS_DIR, f"{video_id}_meta.json")
    with open(meta_path, encoding="utf-8") as f:
        meta = json.load(f)
    video_title = meta.get("title", "")
    video_url   = meta.get("url", "")

    # Entity matching
    matched = {}
    for key, names in entities.items():
        matched[key] = [name for name in names if name in full_text]

    enriched = {
        "video_id":    video_id,
        "video_title": video_title,
        "video_url":   video_url,
        "full_text":   full_text,
        "characters":  matched["characters"],
        "cards":       matched["cards"],
        "relics":      matched["relics"],
        "potions":     matched["potions"],
    }

    with open(enriched_path, "w", encoding="utf-8") as f:
        json.dump(enriched, f, ensure_ascii=False, indent=2)

    print(f"[enrich] {video_id}: chars={len(matched['characters'])}, "
          f"cards={len(matched['cards'])}, relics={len(matched['relics'])}, "
          f"potions={len(matched['potions'])}")
    return enriched


def enrich_all() -> list[dict]:
    entities = load_all_entities()
    print(f"[enrich] Loaded entities: "
          f"chars={len(entities['characters'])}, cards={len(entities['cards'])}, "
          f"relics={len(entities['relics'])}, potions={len(entities['potions'])}")

    video_ids = [
        f.replace("_chunks.json", "")
        for f in os.listdir(TRANSCRIPT_DIR)
        if f.endswith("_chunks.json")
    ]
    video_ids.sort()
    print(f"[enrich] Found {len(video_ids)} videos: {video_ids}")

    results = []
    for vid in video_ids:
        results.append(enrich_video(vid, entities))
    return results


if __name__ == "__main__":
    results = enrich_all()
    print(f"\n[enrich] Done. {len(results)} enriched files written.")
    for r in results:
        print(f"  {r['video_id']}: {r['video_title'][:30]}")
        print(f"    chars={r['characters']}, relics={r['relics']}")
        print(f"    cards={r['cards'][:5]}{'...' if len(r['cards']) > 5 else ''}")
        print(f"    potions={r['potions']}")
