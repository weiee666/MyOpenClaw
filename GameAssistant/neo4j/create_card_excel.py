# ============================================================
# 杀戮尖塔2 卡牌数据 → Excel
# 数据来源：https://sts2.huijiwiki.com/api/rest_v1/namespace/data
# ============================================================
# 生成 2 张表：
#   08_节点_卡牌.xlsx         —— 基础卡牌节点（非升级版）
#   09_关系_卡牌_升级.xlsx    —— Card -[UPGRADES_TO]-> Card
# ============================================================

import json, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = os.path.dirname(__file__)
DATA_FILE  = "/tmp/all_cards.json"

# ── 样式（与 create_excel.py 一致）──────────────────────────
HEADER_FILLS = {
    "node": PatternFill("solid", fgColor="1F4E79"),
    "rel":  PatternFill("solid", fgColor="7B2D2D"),
}
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Microsoft YaHei")
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=True)
THIN        = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
ROW_FILLS = [
    PatternFill("solid", fgColor="EBF3FB"),
    PatternFill("solid", fgColor="FFFFFF"),
]

def make_wb(sheet_name, headers, rows, table_type="node", col_widths=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.row_dimensions[1].height = 28

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILLS[table_type]
        c.alignment = CENTER
        c.border = THIN

    for r_idx, row in enumerate(rows, 2):
        fill = ROW_FILLS[r_idx % 2]
        ws.row_dimensions[r_idx].height = 20
        for c_idx, val in enumerate(row, 1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.fill = fill
            c.border = THIN
            c.alignment = LEFT

    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    else:
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
    return wb

def save(wb, filename):
    path = os.path.join(OUTPUT_DIR, filename)
    wb.save(path)
    print(f"  ✓ {filename}")


# ── 费用中文 → 数字映射 ──────────────────────────────────────
COST_MAP = {"零": 0, "一": 1, "二": 2, "三": 3,
            "四": 4, "五": 5, "六": 6, "X": -1, "无": -2}

# ── 颜色 → 角色 ID 映射 ─────────────────────────────────────
COLOR_TO_CHAR = {
    "红色": "CHAR_001",  # 铁甲战士
    "绿色": "CHAR_002",  # 静默猎手
    "橙色": "CHAR_004",  # 储君（wiki里储君是橙色）
    "紫色": "CHAR_005",  # 亡灵契约师
    "蓝色": "CHAR_003",  # 故障机器人
    "无色": "COMMON",
    "灰色": "COMMON",
    "深灰": "COMMON",
    "黑色": "COMMON",
}

# ============================================================
# 读取数据
# ============================================================
with open(DATA_FILE, encoding="utf-8") as f:
    all_cards = json.load(f)

print(f"原始数据：{len(all_cards)} 条")

# 判断是否为升级版（upgrade 字段值为"已升级"，或 id 以 _upgrade 结尾）
def is_upgraded(card):
    return (
        card.get("upgrade") == "已升级"
        or str(card.get("id", "")).endswith("_upgrade")
    )

base_cards    = [c for c in all_cards if not is_upgraded(c)]
upgrade_cards = [c for c in all_cards if is_upgraded(c)]

print(f"  基础卡牌：{len(base_cards)} 张")
print(f"  升级版本：{len(upgrade_cards)} 张")

# 用 id 建索引，方便做关系表
id_to_card = {c["id"]: c for c in all_cards}

# ============================================================
# 08  节点_卡牌
# ============================================================
card_headers = [
    "card_id",        # 系统 ID（字母）
    "name",           # 中文名
    "color",          # 颜色
    "character_id",   # 对应角色
    "rarity",         # 稀有度
    "type",           # 类型（攻击/技能/能力/状态/诅咒…）
    "cost",           # 费用（中文）
    "cost_num",       # 费用（数字，-1=X，-2=无）
    "description",    # 效果描述
    "upgrade_card_id",# 升级后的卡牌 id
    "has_upgrade",    # 是否有升级版
    "image",          # 图片文件名
    "wiki_page",      # wiki 页面名
]

card_rows = []
for c in base_cards:
    upgrade_id = c.get("upgrade", "")
    has_upg = upgrade_id and upgrade_id != "已升级"
    card_rows.append([
        c.get("id", ""),
        c.get("name", ""),
        c.get("color", ""),
        COLOR_TO_CHAR.get(c.get("color", ""), "COMMON"),
        c.get("rarity", ""),
        c.get("type", ""),
        c.get("cost", ""),
        COST_MAP.get(c.get("cost", ""), ""),
        c.get("description_raw", ""),
        upgrade_id if has_upg else "",
        "TRUE" if has_upg else "FALSE",
        c.get("image", ""),
        c.get("page", "") or "",
    ])

wb = make_wb("卡牌", card_headers, card_rows, "node",
             [24, 16, 8, 12, 10, 10, 8, 10, 40, 24, 12, 28, 18])
save(wb, "08_节点_卡牌.xlsx")


# ============================================================
# 09  关系_卡牌_升级 (Card -[UPGRADES_TO]-> Card)
# ============================================================
rel_headers = [
    "relationship_id",
    "start_id(Card)",      # 基础版 card_id
    "end_id(Card)",        # 升级版 card_id
    "relationship_type",
    "base_name",
    "upgrade_name",
]

rel_rows = []
idx = 1
for c in base_cards:
    upg_id = c.get("upgrade", "")
    if upg_id and upg_id != "已升级" and upg_id in id_to_card:
        upg_card = id_to_card[upg_id]
        rel_rows.append([
            f"RCU_{idx:04d}",
            c["id"],
            upg_id,
            "UPGRADES_TO",
            c.get("name", ""),
            upg_card.get("name", ""),
        ])
        idx += 1

wb = make_wb("卡牌升级", rel_headers, rel_rows, "rel",
             [14, 24, 24, 14, 18, 18])
save(wb, "09_关系_卡牌_升级.xlsx")


# ============================================================
# 统计
# ============================================================
from collections import Counter
color_cnt  = Counter(c.get("color","")  for c in base_cards)
rarity_cnt = Counter(c.get("rarity","") for c in base_cards)
type_cnt   = Counter(c.get("type","")   for c in base_cards)

print()
print("=" * 55)
print("卡牌数据统计")
print("-" * 55)
print(f"基础卡牌总数：{len(base_cards)} 张    升级版：{len(upgrade_cards)} 张")
print()
print("按颜色（角色）分布：")
for k, v in color_cnt.most_common():
    print(f"  {k:6s}: {v} 张")
print()
print("按稀有度分布：")
for k, v in rarity_cnt.most_common():
    print(f"  {k:6s}: {v} 张")
print()
print("按类型分布：")
for k, v in type_cnt.most_common():
    print(f"  {k:8s}: {v} 张")
print()
print(f"有升级版的卡牌：{len(rel_rows)} 张")
print("=" * 55)
