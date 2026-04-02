# ============================================================
# 杀戮尖塔2 Wiki → Neo4j 原始数据 Excel 生成器
# 数据来源：https://sts2.huijiwiki.com/wiki
# ============================================================
# 生成 7 张 Excel 表：4 张节点表 + 3 张关系表
# ============================================================

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUTPUT_DIR = os.path.dirname(__file__)

# ── 样式 ────────────────────────────────────────────────────
HEADER_FILLS = {
    "node":  PatternFill("solid", fgColor="1F4E79"),   # 深蓝 → 节点表
    "rel":   PatternFill("solid", fgColor="7B2D2D"),   # 深红 → 关系表
}
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Microsoft YaHei")
CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)
THIN_BORDER   = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
ROW_FILLS = [
    PatternFill("solid", fgColor="EBF3FB"),
    PatternFill("solid", fgColor="FFFFFF"),
]

def make_wb(sheet_name: str, headers: list[str], rows: list[list],
            table_type: str = "node", col_widths: list[int] | None = None) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.row_dimensions[1].height = 28

    # Header row
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font   = HEADER_FONT
        cell.fill   = HEADER_FILLS[table_type]
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # Data rows
    for r_idx, row in enumerate(rows, 2):
        fill = ROW_FILLS[r_idx % 2]
        ws.row_dimensions[r_idx].height = 20
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill      = fill
            cell.border    = THIN_BORDER
            cell.alignment = LEFT

    # Column widths
    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    else:
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

    return wb

def save(wb, filename):
    path = os.path.join(OUTPUT_DIR, filename)
    wb.save(path)
    print(f"  ✓ {filename}")


# ============================================================
# 01  节点_角色
# ============================================================
char_headers = [
    "character_id", "name", "description", "card_color",
    "is_starter", "origin", "special_mechanic"
]
char_rows = [
    ["CHAR_001", "铁甲战士",   "铁甲军团最后的士兵。用刀剑和烈焰击溃敌人。",     "红色", "TRUE",  "来自杀戮尖塔1", "燃烧(Burn)"],
    ["CHAR_002", "静默猎手",   "一位来自尖塔之外的女猎手，擅长刀刺与毒杀。",     "绿色", "FALSE", "来自杀戮尖塔1", "毒(Poison)"],
    ["CHAR_003", "故障机器人", "不停改造自己的自动机械，使用充能球科技作战。",   "蓝色", "FALSE", "来自杀戮尖塔1", "充能球(Orb)"],
    ["CHAR_004", "储君",       "群星王座的继承人，从宇宙中汲取力量。",           "紫色", "FALSE", "续作新角色",   "辉星(Wis)"],
    ["CHAR_005", "亡灵契约师", "出生在尖塔的巫妖，召唤死灵巨手奥斯提作战。",   "橙色", "FALSE", "续作新角色",   "奥斯提(Ostis)"],
]
wb = make_wb("角色", char_headers, char_rows, "node",
             [14, 14, 36, 10, 12, 16, 18])
save(wb, "01_节点_角色.xlsx")


# ============================================================
# 02  节点_遗物
# ============================================================
relic_headers = [
    "relic_id", "name", "description", "rarity", "character_id", "obtain_source"
]
relic_rows = [
    # 初始遗物（角色专属）
    ["REL_001", "燃烧之血",   "在战斗结束时，回复6点生命。",                     "初始遗物", "CHAR_001", "角色初始携带"],
    ["REL_002", "蛇之戒指",   "在每场战斗开始时，额外抽2张牌。",                  "初始遗物", "CHAR_002", "角色初始携带"],
    ["REL_003", "破损核心",   "在每场战斗开始时，生成1个闪电充能球。",             "初始遗物", "CHAR_003", "角色初始携带"],
    ["REL_004", "天赋君权",   "在每场战斗开始时，获得辉星。",                     "初始遗物", "CHAR_004", "角色初始携带"],
    ["REL_005", "缚魂命匣",   "在你的回合开始时，召唤1点奥斯提。",                "初始遗物", "CHAR_005", "角色初始携带"],
    # 遗物稀有度分类（占位行，待后续补全具体遗物）
    ["REL_T01", "（普通遗物）",  "很容易找到的弱小遗物。",  "普通遗物",  "通用", "击败精英/宝箱/商店"],
    ["REL_T02", "（罕见遗物）",  "比普通遗物更强大的遗物。", "罕见遗物",  "通用", "击败精英/宝箱/商店"],
    ["REL_T03", "（稀有遗物）",  "极为少见的独特遗物。",    "稀有遗物",  "通用", "击败精英/宝箱"],
    ["REL_T04", "（Boss遗物）",  "只在Boss宝箱中出现。",    "Boss遗物",  "通用", "击败Boss"],
    ["REL_T05", "（事件遗物）",  "只能通过事件获得。",      "事件遗物",  "通用", "事件"],
    ["REL_T06", "（先古之民遗物）","只能通过先古之民获得。", "先古之民遗物","通用","先古之民"],
    ["REL_T07", "（商店遗物）",  "只能从商人处购买。",      "商店遗物",  "通用", "商店"],
    ["REL_T08", "头环",         "这是一个头环。",           "其他遗物",  "通用", "收集全部遗物后获得"],
]
wb = make_wb("遗物", relic_headers, relic_rows, "node",
             [12, 16, 36, 14, 12, 20])
save(wb, "02_节点_遗物.xlsx")


# ============================================================
# 03  节点_区域（含阶段信息）
# ============================================================
zone_headers = [
    "zone_id", "zone_name", "stage_number", "stage_name", "description"
]
zone_rows = [
    ["ZONE_001", "密林", 1, "第一阶段", "第一阶段中的密林区域，充满各类植物系与自然系怪物。"],
    ["ZONE_002", "暗港", 1, "第一阶段", "第一阶段中的暗港区域，充满邪教与水生系怪物。"],
    ["ZONE_003", "巢穴", 2, "第二阶段", "第二阶段，充满虫类与巢穴系怪物。"],
    ["ZONE_004", "荣耀", 3, "第三阶段", "第三阶段，充满机械与魔法系精英怪。"],
]
wb = make_wb("区域", zone_headers, zone_rows, "node",
             [12, 10, 14, 14, 40])
save(wb, "03_节点_区域.xlsx")


# ============================================================
# 04  节点_怪物
# ============================================================
monster_headers = [
    "monster_id", "name", "monster_type", "zone_id", "stage_number",
    "is_group", "parent_monster_id", "notes"
]

# 数据：(name, type, zone_id, stage, is_group, parent_id, notes)
raw_monsters = [
    # ── 密林 普通 ──
    ("飞蝇菌子",       "普通", "ZONE_001", 1, False, "",          ""),
    ("劫掠者团伙",     "普通", "ZONE_001", 1, True,  "",          "由多个劫掠者组成"),
    ("劫掠者暴徒",     "普通", "ZONE_001", 1, False, "劫掠者团伙","劫掠者团伙成员"),
    ("劫掠者刺客",     "普通", "ZONE_001", 1, False, "劫掠者团伙","劫掠者团伙成员"),
    ("劫掠者斧手",     "普通", "ZONE_001", 1, False, "劫掠者团伙","劫掠者团伙成员"),
    ("劫掠者弩手",     "普通", "ZONE_001", 1, False, "劫掠者团伙","劫掠者团伙成员"),
    ("劫掠者追踪手",   "普通", "ZONE_001", 1, False, "劫掠者团伙","劫掠者团伙成员"),
    ("雾菇",           "普通", "ZONE_001", 1, True,  "",          "会召唤利齿之眼"),
    ("利齿之眼",       "普通", "ZONE_001", 1, False, "雾菇",      "雾菇召唤物"),
    ("蛮兽",           "普通", "ZONE_001", 1, False, "",          ""),
    ("毛绒伏地虫",     "普通", "ZONE_001", 1, False, "",          ""),
    ("墨宝",           "普通", "ZONE_001", 1, False, "",          ""),
    ("闪光贾克斯果",   "普通", "ZONE_001", 1, False, "",          ""),
    ("蛇行扼杀者",     "普通", "ZONE_001", 1, False, "",          ""),
    ("树叶史莱姆（小）","普通","ZONE_001", 1, False, "",          ""),
    ("树叶史莱姆（中）","普通","ZONE_001", 1, False, "",          ""),
    ("树枝史莱姆（小）","普通","ZONE_001", 1, False, "",          ""),
    ("树枝史莱姆（中）","普通","ZONE_001", 1, False, "",          ""),
    ("缩小甲虫",       "普通", "ZONE_001", 1, False, "",          ""),
    ("藤蔓蹒跚者",     "普通", "ZONE_001", 1, False, "",          ""),
    ("小啃兽",         "普通", "ZONE_001", 1, False, "",          ""),
    ("立柱构造体",     "普通", "ZONE_001", 1, False, "",          "第三阶段也出现"),
    # ── 密林 精英 ──
    ("多尼斯异鸟",     "精英", "ZONE_001", 1, False, "",          ""),
    ("旧日雕像",       "精英", "ZONE_001", 1, False, "",          ""),
    ("异蛙寄生虫",     "精英", "ZONE_001", 1, True,  "",          "会召唤扭动虫"),
    ("扭动虫",         "精英", "ZONE_001", 1, False, "异蛙寄生虫","异蛙寄生虫召唤物"),
    # ── 密林 Boss ──
    ("墨影幻灵",       "Boss", "ZONE_001", 1, False, "",          ""),
    ("同族神官",       "Boss", "ZONE_001", 1, True,  "",          "会召唤同族信徒"),
    ("同族信徒",       "Boss", "ZONE_001", 1, False, "同族神官",  "同族神官召唤物"),
    ("仪式兽",         "Boss", "ZONE_001", 1, False, "",          ""),
    # ── 暗港 普通 ──
    ("蟾蜍蝌蚪",       "普通", "ZONE_002", 1, False, "",          ""),
    ("潮湿邪教徒",     "普通", "ZONE_002", 1, False, "",          ""),
    ("钙化邪教徒",     "普通", "ZONE_002", 1, False, "",          ""),
    ("地精佣兵",       "普通", "ZONE_002", 1, True,  "",          "多地精组合"),
    ("卑鄙地精",       "普通", "ZONE_002", 1, False, "地精佣兵",  "地精佣兵成员"),
    ("胖地精",         "普通", "ZONE_002", 1, False, "地精佣兵",  "地精佣兵成员"),
    ("海洋混混",       "普通", "ZONE_002", 1, False, "",          ""),
    ("化石追踪者",     "普通", "ZONE_002", 1, False, "",          ""),
    ("活雾",           "普通", "ZONE_002", 1, True,  "",          "会生成气态炸弹"),
    ("气态炸弹",       "普通", "ZONE_002", 1, False, "活雾",      "活雾生成物"),
    ("噬尸蛞蝓",       "普通", "ZONE_002", 1, False, "",          ""),
    ("双尾鼠",         "普通", "ZONE_002", 1, False, "",          ""),
    ("下水道蚌",       "普通", "ZONE_002", 1, False, "",          ""),
    ("幽灵船",         "普通", "ZONE_002", 1, False, "",          ""),
    ("淤泥旋螺",       "普通", "ZONE_002", 1, False, "",          ""),
    ("拳击构装体",     "普通", "ZONE_002", 1, False, "",          "第三阶段也出现"),
    # ── 暗港 精英 ──
    ("鬼祟珊瑚群",     "精英", "ZONE_002", 1, False, "",          ""),
    ("骇鳗",           "精英", "ZONE_002", 1, False, "",          ""),
    ("花园幽灵鳗",     "精英", "ZONE_002", 1, False, "",          ""),
    # ── 暗港 Boss ──
    ("乐加维林族母",   "Boss", "ZONE_002", 1, False, "",          ""),
    ("灵魂异鱼",       "Boss", "ZONE_002", 1, False, "",          ""),
    ("瀑布巨兽",       "Boss", "ZONE_002", 1, False, "",          ""),
    # ── 巢穴 普通 ──
    ("地道虫",         "普通", "ZONE_003", 2, False, "",          "特性：埋地"),
    ("棘刺蟾蜍",       "普通", "ZONE_003", 2, False, "",          ""),
    ("啃咬机",         "普通", "ZONE_003", 2, False, "",          ""),
    ("猎人杀手",       "普通", "ZONE_003", 2, False, "",          ""),
    ("胧光怪",         "普通", "ZONE_003", 2, True,  "",          "会召唤寄生惧魔"),
    ("寄生惧魔",       "普通", "ZONE_003", 2, False, "胧光怪",    "胧光怪召唤物"),
    ("盛碗虫",         "普通", "ZONE_003", 2, True,  "",          "有四种形态"),
    ("盛碗虫（卵）",   "普通", "ZONE_003", 2, False, "盛碗虫",    "盛碗虫形态"),
    ("盛碗虫（蜜）",   "普通", "ZONE_003", 2, False, "盛碗虫",    "盛碗虫形态"),
    ("盛碗虫（石）",   "普通", "ZONE_003", 2, False, "盛碗虫",    "盛碗虫形态"),
    ("盛碗虫（丝）",   "普通", "ZONE_003", 2, False, "盛碗虫",    "盛碗虫形态"),
    ("虱虫之祖",       "普通", "ZONE_003", 2, False, "",          ""),
    ("熟睡甲虫",       "普通", "ZONE_003", 2, False, "",          ""),
    ("偷窃草蜢",       "普通", "ZONE_003", 2, False, "",          ""),
    ("外骨骼虫",       "普通", "ZONE_003", 2, False, "",          ""),
    ("异螨",           "普通", "ZONE_003", 2, False, "",          ""),
    ("直飞产卵虫",     "普通", "ZONE_003", 2, True,  "",          "会产卵孵化幼虫"),
    ("结实的卵",       "普通", "ZONE_003", 2, False, "直飞产卵虫","直飞产卵虫产物"),
    ("幼虫",           "普通", "ZONE_003", 2, False, "直飞产卵虫","直飞产卵虫孵化"),
    # ── 巢穴 精英 ──
    ("残杀千足虫",     "精英", "ZONE_003", 2, False, "",          ""),
    ("蜂群术士",       "精英", "ZONE_003", 2, False, "",          ""),
    ("感染棱柱",       "精英", "ZONE_003", 2, False, "",          ""),
    # ── 巢穴 Boss ──
    ("帝皇蟹",         "Boss", "ZONE_003", 2, True,  "",          "附带两个附件"),
    ("碾碎爪",         "Boss", "ZONE_003", 2, False, "帝皇蟹",    "帝皇蟹附件"),
    ("火箭",           "Boss", "ZONE_003", 2, False, "帝皇蟹",    "帝皇蟹附件"),
    ("无厌沙虫",       "Boss", "ZONE_003", 2, False, "",          ""),
    ("知识恶魔",       "Boss", "ZONE_003", 2, False, "",          ""),
    # ── 荣耀 普通 ──
    ("电球头",         "普通", "ZONE_004", 3, False, "",          ""),
    ("高塔炮手",       "普通", "ZONE_004", 3, True,  "",          "携带活体盾"),
    ("活体盾",         "普通", "ZONE_004", 3, False, "高塔炮手",  "高塔炮手的盾牌"),
    ("巨斧机器人",     "普通", "ZONE_004", 3, False, "",          ""),
    ("猫头鹰法官",     "普通", "ZONE_004", 3, False, "",          ""),
    ("虔诚雕刻师",     "普通", "ZONE_004", 3, False, "",          ""),
    ("青蛙骑士",       "普通", "ZONE_004", 3, False, "",          ""),
    ("史莱姆狂战士",   "普通", "ZONE_004", 3, False, "",          ""),
    ("失落与遗忘之物", "普通", "ZONE_004", 3, True,  "",          "两体合一怪"),
    ("失落之物",       "普通", "ZONE_004", 3, False, "失落与遗忘之物","组合成员"),
    ("遗忘之物",       "普通", "ZONE_004", 3, False, "失落与遗忘之物","组合成员"),
    ("咬人卷轴",       "普通", "ZONE_004", 3, False, "",          ""),
    ("战斗好伙伴V1.0", "普通", "ZONE_004", 3, False, "",          ""),
    ("战斗好伙伴V2.0", "普通", "ZONE_004", 3, False, "",          ""),
    ("战斗好伙伴V3.0", "普通", "ZONE_004", 3, False, "",          ""),
    ("组装师",         "普通", "ZONE_004", 3, True,  "",          "会召唤各类机器人"),
    ("守护机器人",     "普通", "ZONE_004", 3, False, "组装师",    "组装师召唤物"),
    ("电击机器人",     "普通", "ZONE_004", 3, False, "组装师",    "组装师召唤物"),
    ("戳刺机器人",     "普通", "ZONE_004", 3, False, "组装师",    "组装师召唤物"),
    ("噪音机器人",     "普通", "ZONE_004", 3, False, "组装师",    "组装师召唤物"),
    # ── 荣耀 精英 ──
    ("机甲骑士",       "精英", "ZONE_004", 3, False, "",          ""),
    ("灵魂枢纽",       "精英", "ZONE_004", 3, False, "",          ""),
    ("骑士团伙",       "精英", "ZONE_004", 3, True,  "",          "三骑士组合"),
    ("连枷骑士",       "精英", "ZONE_004", 3, False, "骑士团伙",  "骑士团伙成员"),
    ("幽灵骑士",       "精英", "ZONE_004", 3, False, "骑士团伙",  "骑士团伙成员"),
    ("魔法骑士",       "精英", "ZONE_004", 3, False, "骑士团伙",  "骑士团伙成员"),
    # ── 荣耀 Boss ──
    ("门扉缔造者",     "Boss", "ZONE_004", 3, True,  "",          ""),
    ("门扉",           "Boss", "ZONE_004", 3, False, "门扉缔造者","门扉缔造者组件"),
    ("女王",           "Boss", "ZONE_004", 3, True,  "",          ""),
    ("火炬头聚合体",   "Boss", "ZONE_004", 3, False, "女王",      "女王组件"),
    ("实验体",         "Boss", "ZONE_004", 3, False, "",          ""),
]

# 构建 name→id 映射（用于关系表）
monster_name_to_id = {}
monster_rows = []
for idx, (name, mtype, zone, stage, is_grp, parent, notes) in enumerate(raw_monsters, 1):
    mid = f"MON_{idx:03d}"
    monster_name_to_id[name] = mid
    monster_rows.append([mid, name, mtype, zone, stage,
                         "TRUE" if is_grp else "FALSE",
                         parent, notes])

# 把 parent_monster_id 从名字换成 id
for row in monster_rows:
    if row[6]:   # parent column
        row[6] = monster_name_to_id.get(row[6], row[6])

wb = make_wb("怪物", monster_headers, monster_rows, "node",
             [12, 18, 10, 12, 14, 10, 16, 30])
save(wb, "04_节点_怪物.xlsx")


# ============================================================
# 05  关系_角色_初始遗物
# ============================================================
rel1_headers = [
    "relationship_id", "start_id(Character)", "end_id(Relic)",
    "relationship_type", "notes"
]
rel1_rows = [
    ["RCR_001", "CHAR_001", "REL_001", "HAS_STARTING_RELIC", "铁甲战士 → 燃烧之血"],
    ["RCR_002", "CHAR_002", "REL_002", "HAS_STARTING_RELIC", "静默猎手 → 蛇之戒指"],
    ["RCR_003", "CHAR_003", "REL_003", "HAS_STARTING_RELIC", "故障机器人 → 破损核心"],
    ["RCR_004", "CHAR_004", "REL_004", "HAS_STARTING_RELIC", "储君 → 天赋君权"],
    ["RCR_005", "CHAR_005", "REL_005", "HAS_STARTING_RELIC", "亡灵契约师 → 缚魂命匣"],
]
wb = make_wb("角色_初始遗物", rel1_headers, rel1_rows, "rel",
             [16, 22, 16, 22, 28])
save(wb, "05_关系_角色_初始遗物.xlsx")


# ============================================================
# 06  关系_怪物_区域
# ============================================================
rel2_headers = [
    "relationship_id", "start_id(Monster)", "end_id(Zone)",
    "relationship_type", "monster_type", "monster_name", "zone_name"
]
rel2_rows = []
for r_idx, row in enumerate(monster_rows, 1):
    mid, name, mtype, zone_id, stage = row[0], row[1], row[2], row[3], row[4]
    zone_name_map = {"ZONE_001":"密林","ZONE_002":"暗港","ZONE_003":"巢穴","ZONE_004":"荣耀"}
    rel2_rows.append([
        f"RMZ_{r_idx:03d}", mid, zone_id,
        "APPEARS_IN", mtype, name, zone_name_map.get(zone_id, zone_id)
    ])
wb = make_wb("怪物_区域", rel2_headers, rel2_rows, "rel",
             [14, 14, 12, 14, 12, 20, 10])
save(wb, "06_关系_怪物_区域.xlsx")


# ============================================================
# 07  关系_怪物_组合（PART_OF）
# ============================================================
rel3_headers = [
    "relationship_id", "start_id(Child Monster)", "end_id(Parent Monster)",
    "relationship_type", "child_name", "parent_name"
]
rel3_rows = []
r_idx = 1
for row in monster_rows:
    mid, name, parent_id = row[0], row[1], row[6]
    if parent_id:
        parent_name = next((r[1] for r in monster_rows if r[0] == parent_id), parent_id)
        rel3_rows.append([
            f"RPO_{r_idx:03d}", mid, parent_id,
            "PART_OF", name, parent_name
        ])
        r_idx += 1

wb = make_wb("怪物_组合", rel3_headers, rel3_rows, "rel",
             [14, 22, 22, 12, 20, 20])
save(wb, "07_关系_怪物_组合.xlsx")


# ============================================================
# Summary
# ============================================================
print()
print("=" * 55)
print("完成！共生成 7 张 Excel 表：")
print("-" * 55)
print("节点表（4张）：")
print("  01_节点_角色.xlsx          →  5 行")
print(f"  02_节点_遗物.xlsx          → {len(relic_rows)} 行")
print(f"  03_节点_区域.xlsx          → {len(zone_rows)} 行")
print(f"  04_节点_怪物.xlsx          → {len(monster_rows)} 行")
print("关系表（3张）：")
print(f"  05_关系_角色_初始遗物.xlsx  → {len(rel1_rows)} 行")
print(f"  06_关系_怪物_区域.xlsx      → {len(rel2_rows)} 行")
print(f"  07_关系_怪物_组合.xlsx      → {len(rel3_rows)} 行")
print("=" * 55)
print(f"输出目录: {OUTPUT_DIR}")
