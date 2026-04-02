# ============================================================
# 杀戮尖塔2 药水 & 遗物数据收集 + Excel 生成（含图片）
# ============================================================

import json, os, re, time, tempfile
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
from playwright.sync_api import sync_playwright, BrowserContext

OUTPUT_DIR = os.path.dirname(__file__)
IMG_POTION = os.path.join(OUTPUT_DIR, "images", "potions")
IMG_RELIC  = os.path.join(OUTPUT_DIR, "images", "relics")
os.makedirs(IMG_POTION, exist_ok=True)
os.makedirs(IMG_RELIC,  exist_ok=True)

THUMB_SIZE = 56
ROW_H_PT   = THUMB_SIZE * 0.75 + 6


def _norm(s): return re.sub(r'[\s_]+', '', s).lower()
def _build_norm_map(img_dir): return {_norm(f): f for f in os.listdir(img_dir)}
def resolve_img(api_name, norm_map, img_dir):
    if not api_name: return None
    real = norm_map.get(_norm(api_name))
    return os.path.join(img_dir, real) if real else None


COLOR_TO_CHAR = {
    "红色": "CHAR_001", "绿色": "CHAR_002", "蓝色": "CHAR_003",
    "橙色": "CHAR_004", "紫色": "CHAR_005",
    "通用": "COMMON",   "无色": "COMMON",
}


# ── 样式工具 ─────────────────────────────────────────────────
def fill(hex_c): return PatternFill("solid", fgColor=hex_c)
H_FILL_NODE = fill("1F4E79")
H_FONT = Font(bold=True, color="FFFFFF", name="Microsoft YaHei", size=10)
D_FONT = Font(name="Microsoft YaHei", size=9)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
THIN   = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"),  bottom=Side(style="thin"))
R_FILL = [fill("EBF3FB"), fill("FFFFFF")]


# ── API helpers ──────────────────────────────────────────────
def fetch_all(ctx: BrowserContext, category: str) -> list:
    items, pg = [], 1
    while True:
        text = None
        for _ in range(3):
            r = ctx.request.get(
                "https://sts2.huijiwiki.com/api/rest_v1/namespace/data",
                params={"filter": '{"category":"' + category + '"}',
                        "pagesize": "100", "page": str(pg)}
            )
            t = r.text()
            if t and t.strip():
                text = t
                break
            time.sleep(1)
        if not text:
            break
        d = json.loads(text)
        items.extend(d.get("_embedded", []))
        total = d.get("_total_pages", 1)
        print(f"    [{category}] 页 {pg}/{total}，累计 {len(items)}", flush=True)
        if pg >= total:
            break
        pg += 1
        time.sleep(0.3)
    return items


def batch_image_urls(ctx: BrowserContext, filenames: list) -> dict:
    url_map = {}
    for i in range(0, len(filenames), 50):
        batch = filenames[i:i+50]
        titles = "|".join(f"File:{fn}" for fn in batch)
        r = ctx.request.get(
            "https://sts2.huijiwiki.com/api.php",
            params={"action": "query", "titles": titles,
                    "prop": "imageinfo", "iiprop": "url", "format": "json"}
        )
        t = r.text()
        if not t:
            continue
        pages = json.loads(t).get("query", {}).get("pages", {})
        for info in pages.values():
            ii = info.get("imageinfo", [])
            if ii:
                fname = info["title"].split(":")[-1]
                url_map[fname] = ii[0]["url"]
        time.sleep(0.2)
    return url_map


def download_images(ctx: BrowserContext, url_map: dict, save_dir: str) -> int:
    ok = 0
    for fname, url in url_map.items():
        dest = os.path.join(save_dir, fname)
        if os.path.exists(dest):
            ok += 1
            continue
        r = ctx.request.get(url)
        if r.status == 200:
            with open(dest, "wb") as f:
                f.write(r.body())
            ok += 1
        time.sleep(0.08)
    return ok


# ── Excel helpers ────────────────────────────────────────────
def setup_ws(ws, headers, col_widths):
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = H_FONT; c.fill = H_FILL_NODE
        c.alignment = CENTER; c.border = THIN
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def prepare_img(img_path: str, tmp_files: list) -> str | None:
    """Convert RGBA→RGB with white background, thumbnail to THUMB_SIZE, save to temp file."""
    if not img_path or not os.path.exists(img_path):
        return None
    try:
        with PILImage.open(img_path) as im:
            if im.mode == "RGBA":
                bg = PILImage.new("RGB", im.size, (255, 255, 255))
                bg.paste(im, mask=im.split()[3])
                img = bg
            else:
                img = im.convert("RGB")
            img.thumbnail((THUMB_SIZE, THUMB_SIZE), PILImage.LANCZOS)
            tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
            img.save(tmp.name, "PNG")
            tmp_files.append(tmp.name)
            return tmp.name
    except Exception:
        return None


def write_data_row(ws, row_idx, values, tmp_img_path=None):
    ws.row_dimensions[row_idx].height = ROW_H_PT
    f = R_FILL[row_idx % 2]
    a = ws.cell(row=row_idx, column=1, value="")
    a.fill = f; a.border = THIN
    for ci, val in enumerate(values, 2):
        c = ws.cell(row=row_idx, column=ci, value=val)
        c.fill = f; c.border = THIN; c.font = D_FONT; c.alignment = LEFT
    if tmp_img_path and os.path.exists(tmp_img_path):
        try:
            with PILImage.open(tmp_img_path) as im:
                w, h = im.size
            xl = XLImage(tmp_img_path)
            xl.width = w; xl.height = h
            ws.add_image(xl, f"A{row_idx}")
        except Exception:
            pass


# ── Excel 生成 ───────────────────────────────────────────────
def make_potion_excel(potions: list):
    headers    = ["图标", "potion_id", "name", "color", "character_id",
                  "tier", "description", "image_file", "wiki_page"]
    col_widths = [10, 24, 14, 8, 12, 10, 50, 28, 16]

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "药水"
    setup_ws(ws, headers, col_widths)

    p_norm = _build_norm_map(IMG_POTION)
    tmp_files = []
    for ri, p in enumerate(potions, 2):
        raw_img = resolve_img(p.get("image", ""), p_norm, IMG_POTION)
        tmp_img = prepare_img(raw_img, tmp_files)
        write_data_row(ws, ri, [
            p.get("id", ""),
            p.get("name", ""),
            p.get("color", ""),
            COLOR_TO_CHAR.get(p.get("color", ""), "COMMON"),
            p.get("tier", ""),
            p.get("description_raw", ""),
            p.get("image", ""),
            p.get("page", "") or "",
        ], tmp_img)

    path = os.path.join(OUTPUT_DIR, "10_节点_药水.xlsx")
    wb.save(path)
    for t in tmp_files:
        try: os.unlink(t)
        except: pass
    print(f"  ✓ 10_节点_药水.xlsx  ({len(potions)} 行)")


def make_relic_excel(relics: list):
    headers    = ["图标", "relic_id", "name", "pool", "character_id",
                  "tier", "description", "flavor", "ancient", "image_file", "wiki_page"]
    col_widths = [10, 28, 16, 8, 12, 10, 50, 30, 10, 30, 16]

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "遗物"
    setup_ws(ws, headers, col_widths)

    r_norm = _build_norm_map(IMG_RELIC)
    tmp_files = []
    for ri, r in enumerate(relics, 2):
        raw_img = resolve_img(r.get("image", ""), r_norm, IMG_RELIC)
        tmp_img = prepare_img(raw_img, tmp_files)
        write_data_row(ws, ri, [
            r.get("id", ""),
            r.get("name", ""),
            r.get("pool", ""),
            COLOR_TO_CHAR.get(r.get("pool", ""), "COMMON"),
            r.get("tier", ""),
            r.get("description_raw", ""),
            r.get("flavor", "") or "",
            "TRUE" if r.get("ancient") else "FALSE",
            r.get("image", ""),
            r.get("page", "") or "",
        ], tmp_img)

    path = os.path.join(OUTPUT_DIR, "02_节点_遗物.xlsx")
    wb.save(path)
    for t in tmp_files:
        try: os.unlink(t)
        except: pass
    print(f"  ✓ 02_节点_遗物.xlsx  ({len(relics)} 行)")


# ── Main ─────────────────────────────────────────────────────
def main():
    print("=" * 55)
    print("杀戮尖塔2 药水 & 遗物数据收集")
    print("=" * 55)

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        ctx = browser.new_context(
            locale="zh-CN",
            user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                       "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
        )
        page = ctx.new_page()
        page.goto("https://sts2.huijiwiki.com/wiki/%E9%A6%96%E9%A1%B5", timeout=30000)
        page.wait_for_timeout(5000)
        print(f"Wiki 已连接：{page.title()}")

        # 1. 拉取数据
        print("\n[1/4] 拉取药水数据...")
        potions = fetch_all(ctx, "potion")
        print(f"  药水共 {len(potions)} 条")

        print("\n[2/4] 拉取遗物数据...")
        relics = fetch_all(ctx, "relic")
        print(f"  遗物共 {len(relics)} 条")

        # 2. 获取图片 URL
        print("\n[3/4] 获取图片 URL...")
        p_imgs = list({p["image"] for p in potions if p.get("image")})
        r_imgs = list({r["image"] for r in relics  if r.get("image")})
        p_url_map = batch_image_urls(ctx, p_imgs)
        r_url_map = batch_image_urls(ctx, r_imgs)
        print(f"  药水图片 URL：{len(p_url_map)}/{len(p_imgs)}")
        print(f"  遗物图片 URL：{len(r_url_map)}/{len(r_imgs)}")

        # 3. 下载图片
        print("\n[4/4] 下载图片...")
        ok_p = download_images(ctx, p_url_map, IMG_POTION)
        ok_r = download_images(ctx, r_url_map, IMG_RELIC)
        print(f"  药水图片：{ok_p}/{len(p_url_map)}")
        print(f"  遗物图片：{ok_r}/{len(r_url_map)}")

        browser.close()

    # 4. 生成 Excel
    print("\n生成 Excel...")
    make_potion_excel(potions)
    make_relic_excel(relics)

    # 统计
    from collections import Counter
    print()
    print("=" * 55)
    print(f"药水 {len(potions)} 条：", dict(Counter(p["tier"] for p in potions)))
    print(f"遗物 {len(relics)} 条：", dict(Counter(r["tier"] for r in relics)))
    print(f"遗物按角色：", dict(Counter(r["pool"] for r in relics)))
    print("=" * 55)


if __name__ == "__main__":
    main()
