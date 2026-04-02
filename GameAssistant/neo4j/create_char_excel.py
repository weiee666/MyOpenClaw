# ============================================================
# 杀戮尖塔2 角色节点表（含角色立绘）
# 重新生成 01_节点_角色.xlsx
# ============================================================

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

OUTPUT_DIR = os.path.dirname(__file__)
IMG_DIR    = os.path.join(OUTPUT_DIR, "images")

# ── 样式 ────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Microsoft YaHei", size=11)
CELL_FONT   = Font(name="Microsoft YaHei", size=10)
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=True)
THIN        = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
ROW_FILLS = [
    PatternFill("solid", fgColor="EBF3FB"),
    PatternFill("solid", fgColor="FFFFFF"),
]

# ── 角色数据 ─────────────────────────────────────────────────
CHARACTERS = [
    {
        "character_id":    "CHAR_001",
        "name":            "铁甲战士",
        "name_en":         "Ironclad",
        "description":     "铁甲军团最后的士兵。并非出自自身的意愿，用刀剑和烈焰击溃敌人们。",
        "card_color":      "红色",
        "is_starter":      "TRUE",
        "origin":          "来自杀戮尖塔1",
        "special_mechanic":"燃烧(Burn)·力量(Strength)",
        "starting_relic":  "燃烧之血",
        "relic_effect":    "在战斗结束时，回复6点生命。",
        "unlock_condition":"游戏初始解锁",
        "image_file":      "Char_select_ironclad.png",
    },
    {
        "character_id":    "CHAR_002",
        "name":            "静默猎手",
        "name_en":         "Silent",
        "description":     "一位来自尖塔之外的女猎手。随时准备刀刺与毒杀任何拦路者。",
        "card_color":      "绿色",
        "is_starter":      "FALSE",
        "origin":          "来自杀戮尖塔1",
        "special_mechanic":"毒(Poison)·飞刀(Shiv)",
        "starting_relic":  "蛇之戒指",
        "relic_effect":    "在每场战斗开始时，额外抽2张牌。",
        "unlock_condition":"随游戏进程解锁",
        "image_file":      "Char_select_silent.png",
    },
    {
        "character_id":    "CHAR_003",
        "name":            "故障机器人",
        "name_en":         "Defect",
        "description":     "一具不停改造自己来生存下去的自动机械。使用充能球科技作战。",
        "card_color":      "蓝色",
        "is_starter":      "FALSE",
        "origin":          "来自杀戮尖塔1",
        "special_mechanic":"充能球(Orb)·专注(Focus)",
        "starting_relic":  "破损核心",
        "relic_effect":    "在每场战斗开始时，生成1个闪电充能球。",
        "unlock_condition":"随游戏进程解锁",
        "image_file":      "Char_select_defect.png",
    },
    {
        "character_id":    "CHAR_004",
        "name":            "储君",
        "name_en":         "Regent",
        "description":     "群星王座的继承人。拥有宇宙的力量，总是让仆从们去做各种事情。",
        "card_color":      "橙色",
        "is_starter":      "FALSE",
        "origin":          "续作新角色",
        "special_mechanic":"辉星(Wis)·仆从(Minion)",
        "starting_relic":  "天赋君权",
        "relic_effect":    "在每场战斗开始时，获得辉星。",
        "unlock_condition":"随游戏进程解锁",
        "image_file":      "Char_select_regent.png",
    },
    {
        "character_id":    "CHAR_005",
        "name":            "亡灵契约师",
        "name_en":         "Necrobinder",
        "description":     "一位出生在尖塔的巫妖，正在寻求复仇。召唤死灵巨手奥斯提作战。",
        "card_color":      "紫色",
        "is_starter":      "FALSE",
        "origin":          "续作新角色",
        "special_mechanic":"奥斯提(Ostis)·墓碑(Tombstone)",
        "starting_relic":  "缚魂命匣",
        "relic_effect":    "在你的回合开始时，召唤1点奥斯提。",
        "unlock_condition":"随游戏进程解锁",
        "image_file":      "Char_select_necrobinder.png",
    },
]

HEADERS = [
    "角色立绘",         # A - 图片列
    "character_id",    # B
    "name",            # C
    "name_en",         # D
    "card_color",      # E
    "is_starter",      # F
    "origin",          # G
    "special_mechanic",# H
    "starting_relic",  # I
    "relic_effect",    # J
    "unlock_condition",# K
    "description",     # L
    "image_file",      # M
]

# 目标行高（像素 → Excel 单位约 0.75）
ROW_H_PX  = 160     # 每行高度（像素）
ROW_H_PT  = ROW_H_PX * 0.75
IMG_H_PX  = 150     # 图片显示高度

# ── 创建工作簿 ───────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "角色"

# 列宽
COL_WIDTHS = [22, 12, 10, 14, 10, 10, 14, 24, 14, 28, 18, 36, 28]
for i, w in enumerate(COL_WIDTHS, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# 表头行
ws.row_dimensions[1].height = 28
for col, h in enumerate(HEADERS, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font      = HEADER_FONT
    c.fill      = HEADER_FILL
    c.alignment = CENTER
    c.border    = THIN

# 数据行
for row_idx, char in enumerate(CHARACTERS, 2):
    ws.row_dimensions[row_idx].height = ROW_H_PT

    fill = ROW_FILLS[row_idx % 2]

    # 文字列（B 开始，跳过 A 图片列）
    data = [
        char["character_id"],
        char["name"],
        char["name_en"],
        char["card_color"],
        char["is_starter"],
        char["origin"],
        char["special_mechanic"],
        char["starting_relic"],
        char["relic_effect"],
        char["unlock_condition"],
        char["description"],
        char["image_file"],
    ]
    # A 列留给图片，背景色一致
    a_cell = ws.cell(row=row_idx, column=1, value="")
    a_cell.fill   = fill
    a_cell.border = THIN

    for col_idx, val in enumerate(data, 2):
        c = ws.cell(row=row_idx, column=col_idx, value=val)
        c.fill      = fill
        c.border    = THIN
        c.font      = CELL_FONT
        c.alignment = LEFT if col_idx > 2 else CENTER

    # ── 嵌入角色图片 ─────────────────────────────────────────
    img_path = os.path.join(IMG_DIR, char["image_file"])
    if os.path.exists(img_path):
        # 用 PIL 计算等比缩放宽度
        with PILImage.open(img_path) as pil_img:
            orig_w, orig_h = pil_img.size
        scale   = IMG_H_PX / orig_h
        disp_w  = int(orig_w * scale)

        xl_img = XLImage(img_path)
        xl_img.height = IMG_H_PX
        xl_img.width  = disp_w

        # 锚定到 A 列该行（行索引从 1 开始，openpyxl cell 坐标）
        cell_addr = f"A{row_idx}"
        ws.add_image(xl_img, cell_addr)
        print(f"  嵌入图片：{char['name']} ({orig_w}x{orig_h} → {disp_w}x{IMG_H_PX})")
    else:
        print(f"  ⚠ 图片不存在：{img_path}")

# ── 冻结首行 ────────────────────────────────────────────────
ws.freeze_panes = "A2"

# ── 保存 ────────────────────────────────────────────────────
out_path = os.path.join(OUTPUT_DIR, "01_节点_角色.xlsx")
wb.save(out_path)
print(f"\n✓ 已保存：01_节点_角色.xlsx（{len(CHARACTERS)} 行，含角色立绘）")
